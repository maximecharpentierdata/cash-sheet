import io
from openpyxl import load_workbook
import datetime
import calendar
from openpyxl.styles import PatternFill, Border, Side

from jours_feries_france import JoursFeries


MONTHS = [
    "Janvier",
    "Février",
    "Mars",
    "Avril",
    "Mai",
    "Juin",
    "Juillet",
    "Août",
    "Septembre",
    "Octobre",
    "Novembre",
    "Décembre",
]
MONTHS = {MONTHS[i - 1]: i for i in range(1, len(MONTHS) + 1)}

DAYS = {
    0: "Lundi",
    1: "Mardi",
    2: "Mercredi",
    3: "Jeudi",
    4: "Vendredi",
    5: "Samedi",
    6: "Dimanche",
}


SUNDAY_FILL = PatternFill(
    start_color="AEAAAA", end_color="AEAAAA", fill_type="solid"
)

OFF_DAY_FILL = PatternFill(
    start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"
)

CLASSIC_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

EUR_CURRENCY_FORMAT = "#,##0.00 €"


def prepare_month(ws, year, off_days, month, last_cell=None):
    ws.title = month
    ws["C4"].value = ws["C4"].value.format(month=month.upper(), year=year)
    month_days = list(
        filter(
            lambda elem: elem[0] > 0,
            calendar.Calendar().itermonthdays2(year, MONTHS[month]),
        )
    )

    if last_cell:
        ws["G6"].value = last_cell
        ws["G6"].number_format = EUR_CURRENCY_FORMAT

    ws["G10"].value = "=G6+E10-D10-F10"
    ws["G10"].number_format = EUR_CURRENCY_FORMAT

    for i, day in enumerate(month_days):
        ws[f"A{10+i}"].value = DAYS[day[1]][0]
        ws[f"B{10+i}"].value = day[0]

        for cell in ws[f"D{10+i}:G{10+i}"][0]:
            cell.number_format = EUR_CURRENCY_FORMAT

        if day[1] == 6:
            for cell in ws[f"A{10+i}:G{10+i}"][0]:
                cell.fill = SUNDAY_FILL
        if datetime.date(year, MONTHS[month], day[0]) in off_days:
            for cell in ws[f"A{10+i}:G{10+i}"][0]:
                cell.fill = OFF_DAY_FILL
        for cell in ws[f"A{10+i}:G{10+i}"][0]:
            cell.border = CLASSIC_BORDER
        if i > 0:
            ws[f"G{10+i}"].value = f"=G{9+i}+E{10+i}-D{10+i}-F{10+i}"

    last_cell = f"={month}!G{10+i}"
    return last_cell


def make_sheet(year, off_days):
    workbook = load_workbook("base.xlsx")
    ref_ws = workbook.active

    last_cell = None
    for month in MONTHS:
        ws = workbook.copy_worksheet(ref_ws)
        last_cell = prepare_month(ws, year, off_days, month, last_cell)

    workbook.remove(ref_ws)

    file = io.BytesIO()

    workbook.save(file)
    return file
